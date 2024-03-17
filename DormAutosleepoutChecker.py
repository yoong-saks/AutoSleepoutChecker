import tkinter as tk
import openpyxl
from tkinter import filedialog
import tkinter.messagebox as messagebox
from datetime import datetime, timedelta
from tkinter import ttk
import pyperclip
import os
import win32com.client as win32
import shutil

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.main_excel_file = ''
        self.sub_excel_file = ''
        self.date = datetime.today().strftime('%Y/%m/%d')
        self.create_widgets()

    def create_widgets(self):
        self.master.title("외박 자동화푸로그램 by 정호성")

        self.windows_usr_name = os.path.expanduser('~')
        backup_dir_path = f"{self.windows_usr_name}\\Desktop\\자동엑셀백업"
        
        if not os.path.exists(backup_dir_path):
            os.mkdir(backup_dir_path)
        

        self.filename = ".autoexcel"
        self.filepath = os.path.join(backup_dir_path, self.filename)
        
        self.hogwan = ttk.Combobox(values=["입지1호관","입지2호관","입지3호관","입지5호관","향토생활관(남)","신애1호관","신애2호관","신애3호관","신애5호관","신애7호관","항토생활관(여)"])
        if os.path.exists(self.filepath):
            with open(self.filepath, "r") as f:
                self.last_selected_value = f.read().strip()
                if self.last_selected_value in self.hogwan["value"]:
                    self.hogwan.set(self.last_selected_value)
        def on_hogwan_select(event):
            with open(self.filepath, "w") as f:
                f.write(self.hogwan.get())
                
        self.hogwan.bind("<<ComboboxSelected>>", on_hogwan_select)

        self.hogwan.pack(side="top")

    
                

        self.main_excel_label = tk.Label(self.master, text="공유폴더 엑셀파일: " + self.main_excel_file)
        self.main_excel_label.pack()

        self.filepathname = '.path_autoexcel'
        self.filepathnamepath = os.path.join(backup_dir_path, self.filepathname)
        if os.path.exists(self.filepathnamepath):
            with open(self.filepathnamepath, "r") as f:
                self.last_selected_path = f.read()
                self.main_excel_file = self.last_selected_path
                self.main_excel_label.config(text="공유폴더 엑셀파일: " + self.last_selected_path)

        self.main_excel_button = tk.Button(self.master, text="Browse", command=self.browse_main_excel)
        self.main_excel_button.pack()

        self.sub_excel_label = tk.Label(self.master, text="외박자 엑셀파일: " + self.sub_excel_file)
        self.sub_excel_label.pack()

        self.sub_excel_button = tk.Button(self.master, text="Browse", command=self.browse_sub_excel)
        self.sub_excel_button.pack()

        self.date_label = tk.Label(self.master, text="날짜: ")
        self.date_label.pack(side="left")
        self.date_entry = tk.Entry(self.master)
        self.date_entry.pack(side="left")
        self.date_entry.insert(0, self.date)

        self.minus_button = tk.Button(self.master, text="-1일", command=self.minus_day)
        self.minus_button.pack(side="left")

        self.plus_button = tk.Button(self.master, text="+1일", command=self.plus_day)
        self.plus_button.pack(side="left")

        self.start_button = tk.Button(self.master, text="프로그램 시작", command=self.start_program)
        self.start_button.pack()

    def browse_main_excel(self):
        self.main_excel_file = filedialog.askopenfilename()
        self.main_excel_label.config(text="공유폴더 엑셀파일: " + self.main_excel_file)

        with open(self.filepathnamepath, "w") as f:
            f.write(self.main_excel_file)

        

    def browse_sub_excel(self):
        self.sub_excel_file = filedialog.askopenfilename()
        self.sub_excel_label.config(text="외박자 엑셀파일: " + self.sub_excel_file)

    def minus_day(self):
        self.date = (datetime.strptime(self.date, '%Y/%m/%d') - timedelta(days=1)).strftime('%Y/%m/%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.date)

    def plus_day(self):
        self.date = (datetime.strptime(self.date, '%Y/%m/%d') + timedelta(days=1)).strftime('%Y/%m/%d')
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, self.date)


        
    def start_program(self):

        def xls_to_xlsx(file_path):
            if file_path.endswith('.xls'):
                now_dir = os.path.abspath('./')
                
                excel = win32.Dispatch('Excel.Application')
                wb = excel.Workbooks.Open(file_path)
                new_file_path = os.path.splitext(file_path)[0] + '.xlsx'
                new_file_path = new_file_path.replace('/', '\\')
                
                wb.SaveAs(new_file_path, FileFormat=51)
                wb.Close()
                excel.Application.Quit()
                os.remove(file_path)
                return new_file_path
            else:
                return file_path

        # 날짜형식
        selected_date_obj = datetime.strptime(self.date, '%Y/%m/%d')
       
        month = selected_date_obj.month
        day = selected_date_obj.day
        
        # 백업 //중요..
        backup_dir_path = f"{self.windows_usr_name}\\Desktop\\자동엑셀백업"
        backup_name = f'{month}월 {day}일.xlsx'
        copy_file_path = os.path.join(backup_dir_path, backup_name)

        if os.path.exists(copy_file_path):
            suffix = 1
            while True:
                new_file_name = f"{backup_name[:-5]} ({suffix}).xlsx"
                new_thisfile_path = os.path.join(backup_dir_path, new_file_name)
                if os.path.exists(new_thisfile_path):
                    suffix += 1
                else:
                    copy_file_path = new_thisfile_path
                    break
                
        shutil.copy(self.main_excel_file,copy_file_path)
        

        
        # open main Excel file
        self.wb_main = openpyxl.load_workbook(self.main_excel_file)
        if "인원관리" in self.wb_main.sheetnames:
            self.ws_main = self.wb_main['인원관리']
            
        elif f"{month}월" in self.wb_main.sheetnames:
            self.ws_main = self.wb_main[f'{month}월']
            
        self.ws_main2 = self.wb_main['사생명단']

        
        # open sub Excel file
        converted_sub_excel_file = xls_to_xlsx(self.sub_excel_file)
        self.wb_sub = openpyxl.load_workbook(converted_sub_excel_file)
        self.ws_sub = self.wb_sub.active


        # 사생명단과 인원관리의 '학번' 내용이 들어가있는 행 찾기
        def myrow_return(mysheet):
            for row in mysheet.iter_rows():
                for cell in row:
                    if cell.value == '학번':
                        return(cell.row)

        sasang_row = myrow_return(self.ws_main2) + 1
        inwon_row = myrow_return(self.ws_main) + 1

        # 오늘에 해당하는 열 찾기
        date_column = None
        for column in range(1, self.ws_main.max_column + 1):
            cell_value = self.ws_main.cell(row=1, column=column).value
            if cell_value:
                if cell_value == f"{month}월 {day}일" or cell_value == f"{month}월{day}일" or cell_value == selected_date_obj.strftime("%m월 %d일") or cell_value == selected_date_obj.strftime("%m월%d일"):  # 생성한 문자열과 비교
                    date_column = column
                    break

                    
        # 서브 엑셀파일에서 해당하는 학생의 학번 찾기
        for row in range(sasang_row, self.ws_sub.max_row + 2):
            student_num = self.ws_sub.cell(row=row-1, column=4).value
            my_sasang_row = sasang_row
            for main_row in range(inwon_row, self.ws_main.max_row + 1):
                main_student_num = self.ws_main2.cell(row=my_sasang_row, column=4).value
                my_sasang_row += 1
                if student_num == main_student_num:
                    cell_value = self.ws_main.cell(row=main_row, column=date_column).value
                    if not cell_value:
                        self.ws_main.cell(row=main_row, column=date_column).value = "외박"
                        
        my_sasang_row = sasang_row
        for main_row in range(inwon_row, self.ws_main.max_row + 1):
            main_student_num = self.ws_main2.cell(row=my_sasang_row, column=4).value
            my_sasang_row += 1
            if not main_student_num:
                continue 
            cell_value = self.ws_main.cell(row=main_row, column=date_column).value
            
            if cell_value == "외박":
                continue 
            if not cell_value:
                self.ws_main.cell(row=main_row, column=date_column).value = "재실"

                    
        # 보고용
        num_students = 0
        num_overnight = 0
        num_occupancy = 0

        my_sasang_row = sasang_row
        for row in range(inwon_row, self.ws_main.max_row + 1):
            student_num = self.ws_main2.cell(row=my_sasang_row, column=4).value
            cell_value = self.ws_main.cell(row=row, column=date_column).value
            my_sasang_row +=1
    
            if student_num:
                num_students += 1
        
                if cell_value == '외박':
                    num_overnight += 1
                elif cell_value == '재실':
                    num_occupancy += 1

        # self.ws_main.cell(row=2, column=date_column).value = num_students
        # self.ws_main.cell(row=3, column=date_column).value = num_overnight
        # self.ws_main.cell(row=5, column=date_column).value = num_occupancy

        self.wb_main.save(self.main_excel_file)

        selected_hogwan = self.hogwan.get()
        bogo_str =  f"{month}월{day}일\n{selected_hogwan}\n총원: {num_students}명\n외박: {num_overnight}명\n재실: {num_occupancy}명"
        pyperclip.copy(bogo_str)
        messagebox.showinfo("진행 완료 됐따~~", "엑셀은 내가 다 수정했다굿.\n보고상황이 복사되었으니 카톡창에 붙이셔")

root = tk.Tk()
app = Application(master=root)
app.mainloop()
